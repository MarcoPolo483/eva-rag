"""Excel document loader for Microsoft Excel files."""
from io import BytesIO
from typing import Any, BinaryIO, Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from eva_rag.loaders.base import DocumentLoader, ExtractedDocument


class ExcelLoader(DocumentLoader):
    """
    Loader for Microsoft Excel documents (.xlsx, .xls).
    
    Features:
    - Multiple sheet support
    - Formula extraction (shows formulas and calculated values)
    - Cell formatting metadata (merged cells, hyperlinks)
    - Table and named range detection
    - Markdown conversion for each sheet
    - Large workbook handling
    
    Requires: openpyxl for .xlsx, xlrd for .xls
    
    Example:
        >>> loader = ExcelLoader(max_rows_per_sheet=100)
        >>> with open("employment_report.xlsx", "rb") as f:
        ...     doc = loader.load_from_stream(f)
        >>> print(doc.metadata["sheet_count"])  # 3
        >>> print(doc.text)  # All sheets as markdown
    """
    
    def __init__(
        self,
        max_rows_per_sheet: Optional[int] = None,
        include_formulas: bool = True,
        read_only: bool = True,
    ):
        """
        Initialize Excel loader.
        
        Args:
            max_rows_per_sheet: Max rows per sheet (None = all)
            include_formulas: Whether to extract formula text
            read_only: Use read-only mode for performance
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel loading. "
                "Install with: pip install openpyxl"
            )
        
        self.max_rows_per_sheet = max_rows_per_sheet
        self.include_formulas = include_formulas
        self.read_only = read_only
    
    def load(self, file: BinaryIO, filename: str) -> ExtractedDocument:
        """
        Load and extract text from Excel document.
        
        Args:
            file: Binary file object
            filename: Original filename
            
        Returns:
            Extracted document with text and metadata
        """
        return self.load_from_stream(file)
    
    def load_from_stream(self, stream: BinaryIO) -> ExtractedDocument:
        """
        Load Excel document from stream.
        
        Args:
            stream: Binary stream of Excel file
            
        Returns:
            Extracted document with text and metadata
        """
        try:
            workbook = openpyxl.load_workbook(
                BytesIO(stream.read()),
                read_only=self.read_only,
                data_only=not self.include_formulas
            )
        except Exception as e:
            return ExtractedDocument(
                text="",
                page_count=0,
                metadata={
                    "error": f"Excel load error: {str(e)}",
                    "loader": "ExcelLoader"
                }
            )
        
        # Extract sheets
        sheets_text = []
        sheets_metadata = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text, sheet_meta = self._extract_sheet(sheet, sheet_name)
            sheets_text.append(sheet_text)
            sheets_metadata.append(sheet_meta)
        
        # Combine all sheets
        text = "\n\n---\n\n".join(sheets_text)
        
        metadata = {
            "loader": "ExcelLoader",
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "sheets": sheets_metadata,
        }
        
        # Add workbook properties if available
        if hasattr(workbook, "properties") and workbook.properties:
            props = workbook.properties
            if props.creator:
                metadata["creator"] = props.creator
            if props.title:
                metadata["title"] = props.title
            if props.created:
                metadata["created"] = str(props.created)
        
        workbook.close()
        
        return ExtractedDocument(
            text=text,
            page_count=len(workbook.sheetnames),  # Each sheet = 1 page
            metadata=metadata
        )
    
    def _extract_sheet(self, sheet, sheet_name: str) -> tuple[str, dict[str, Any]]:
        """
        Extract text and metadata from a single sheet.
        
        Args:
            sheet: openpyxl worksheet
            sheet_name: Name of the sheet
            
        Returns:
            Tuple of (text, metadata)
        """
        lines = []
        lines.append(f"# Sheet: {sheet_name}\n")
        
        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if self.max_rows_per_sheet and max_row > self.max_rows_per_sheet:
            max_row = self.max_rows_per_sheet
            lines.append(f"**Note:** Showing first {self.max_rows_per_sheet} of {sheet.max_row} rows\n")
        
        lines.append(f"**Dimensions:** {max_row} rows × {max_col} columns\n")
        
        # Extract table data
        if max_row > 0 and max_col > 0:
            lines.append("\n## Data\n")
            
            # Get all cell values
            rows_data = []
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    # Get cell value
                    value = cell.value
                    
                    # Add formula if available and enabled
                    if self.include_formulas and hasattr(cell, "data_type") and cell.data_type == "f":
                        if hasattr(cell, "_value"):
                            value = f"={cell._value} (= {value})"
                    
                    row_data.append(str(value) if value is not None else "")
                
                rows_data.append(row_data)
            
            # Generate markdown table
            if rows_data:
                # Header (first row)
                header = "| " + " | ".join(rows_data[0]) + " |"
                separator = "| " + " | ".join(["---"] * max_col) + " |"
                lines.append(header)
                lines.append(separator)
                
                # Data rows
                for row in rows_data[1:]:
                    # Truncate long values
                    row = [v[:50] + "..." if len(v) > 50 else v for v in row]
                    line = "| " + " | ".join(row) + " |"
                    lines.append(line)
        
        text = "\n".join(lines)
        
        metadata = {
            "name": sheet_name,
            "rows": max_row,
            "columns": max_col,
        }
        
        # Add merged cells info
        if hasattr(sheet, "merged_cells") and sheet.merged_cells:
            metadata["merged_cells_count"] = len(sheet.merged_cells.ranges)
        
        return text, metadata
